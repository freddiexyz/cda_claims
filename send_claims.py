from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl import load_workbook, Workbook
from contextlib import contextmanager
from reportlab.pdfgen import canvas
from OpenDental import OpenDental
from docxtpl import DocxTemplate
from sc_variables import *
from datetime import date
from time import sleep
# import acc_automater
# import tkinter.ttk as ttk #stuff
import tkinter as tk
import re
import os

"""
TODO:
Send ACC better
Tidy up all the redundancy but also don't break anything
"""

@contextmanager
def in_dir(path):
    saved = os.getcwd()
    try:
        os.chdir(path)
    except FileNotFoundError:
        os.makedirs(path)
        os.chdir(path)

    yield

    os.chdir(saved)


class SendClaims(OpenDental):
    '''Automatically send claims from OpenDental database'''

    def sql(self, *a, **kw):
        return self.curs.execute(*a, **kw)


    def get_patients_SDSC(self):
        '''Gets patients with SDSC claims waiting to send, returns a tuple of dictionaries'''
        fields = ('claimnum', 'birthdate', 'lname', 'fname', 'address', 'city', 'nhi', 'gender', 'school', 'subid')

        self.sql(SDSC_get_patients)
        patients = tuple({field : value for field, value in zip(fields, values)} for values in self.curs.fetchall())

        for patient in patients:
            if patient['birthdate']:
                patient['birthdate'] = patient['birthdate'].strftime("%d%m%Y")
            patient['city2'] = patient['city'] #required for the form template                

        return patients


    def get_patients_OHSA(self):
        ohsa_fields = ('claimnum', 'birthdate', 'lname', 'fname', 'nhi', 'gender', 'school')

        self.sql(OHSA_get_patients)
        patients = tuple({field : value for field, value in zip(ohsa_fields, values)} for values in self.curs.fetchall())

        for patient in patients:
            if patient['birthdate']:
                patient['birthdate'] = patient['birthdate'].strftime("%d%m%Y")

        return patients


    def get_patients_PA(self):
        fields = ('claimnum', 'birthdate', 'lname', 'fname', 'address', 'city', 'nhi', 'gender', 'school', 'subid', 'pa_num')

        self.sql(PA_SDSC_get_patients)
        patients = tuple({field : value for field, value in zip(fields, values)} for values in self.curs.fetchall())

        for patient in patients:
            if patient['birthdate']:
                patient['birthdate'] = patient['birthdate'].strftime("%d%m%Y")
            patient['city2'] = patient['city']                

        return patients


    def get_patients_OHSA_PA(self):
        ohsa_fields = ('claimnum', 'birthdate', 'lname', 'fname', 'nhi', 'gender', 'school', 'pa_num')

        self.sql(OHSA_PA_get_patients)
        patients = tuple({field : value for field, value in zip(ohsa_fields, values)} for values in self.curs.fetchall())

        for patient in patients:
            if patient['birthdate']:
                patient['birthdate'] = patient['birthdate'].strftime("%d%m%Y")

        return patients



    def get_procedures(self, patient):
        self.sql(get_procedures.format(patient['claimnum']))
        result = self.curs.fetchall()

        keys = ("qty", "date", "fee", "tooth", "code")
        claims = tuple({key : value for key, value in zip(keys, values)} for values in result)

        for claim in claims:
            claim['date'] = claim['date'].strftime("%d.%m.%y")

            claim['qty'] = f"{claim['qty']}"
            claim['fee'] = f"{claim['fee']:.2f}"

            if claim['tooth']:
                _teeth = claim['tooth'].split(',')
                claim['tooth'] = ','.join(teeth[_tooth.strip()] for _tooth in _teeth if _tooth.strip().isalnum())

        return claims


    def get_procedures_PA(self, patient):
        self.sql(PA_get_procedures.format(patient['claimnum']))
        result = self.curs.fetchall()

        keys = ("qty", "date", "fee", "tooth", "code")
        claims = tuple({key : value for key, value in zip(keys, values)} for values in result)

        for claim in claims:
            claim['date'] = claim['date'].strftime("%d.%m.%y")

            claim['qty'] = f"{claim['qty']}"
            claim['fee'] = f"{claim['fee']:.2f}"

            if claim['tooth']:
                _teeth = claim['tooth'].split(',')
                claim['tooth'] = ','.join(teeth[_tooth.strip()] for _tooth in _teeth if _tooth.strip().isalnum())

        return claims


    def get_procedures_OHSA(self, patient):
        self.sql(get_procedures.format(patient['claimnum']))
        result = self.curs.fetchall()

        keys = ("qty", "date", "fee", "tooth", "code")
        claims = tuple({key : value for key, value in zip(keys, values)} for values in result if 'DBCON1' not in values)

        for claim in claims:
            if re.fullmatch("^DBCOM[1-3]$", claim['code']):
                claim['date'] = claim['date'].strftime("%d%m%Y")
                claim['code'] = 'CON1'
            else:
                claim['date'] = claim['date'].strftime("%d%m%y")

            claim['qty'] = f"{claim['qty']}"

            if claim['fee']:
                claim['fee'] = f"{claim['fee']:.2f}"
            else:
                claim['fee'] = ''

            if claim['tooth']:
                _teeth = claim['tooth'].split(',')
                claim['tooth'] = ','.join(teeth[_tooth.strip()] for _tooth in _teeth if _tooth.strip().isalnum())

        return claims


    def get_procedures_OHSA_PA(self, patient):
        self.sql(get_procedures.format(patient['claimnum']))
        result = self.curs.fetchall()

        keys = ("qty", "date", "fee", "tooth", "code")
        claims = tuple({key : value for key, value in zip(keys, values)} for values in result if 'DBCON1' not in values)

        for claim in claims:
            claim['date'] = claim['date'].strftime("%d%m%y")

            claim['qty'] = f"{claim['qty']}"

            if claim['fee']:
                claim['fee'] = f"{claim['fee']:.2f}"
            else:
                claim['fee'] = ''

            if claim['tooth']:
                _teeth = claim['tooth'].split(',')
                claim['tooth'] = ','.join(teeth[_tooth.strip()] for _tooth in _teeth if _tooth.strip().isalnum())

        return claims


    @staticmethod
    def nhi_checksum(nhi):
        '''Uses the check digit to verify that NHI is valid'''
        alpha = 'ABCDEFGHJKLMNPQRSTUVWXYZ' #does not contain 'I' and 'O'

        nhi = str(nhi).upper().strip()
        if not re.fullmatch('^[A-H|J-N|P-Z]{3}[\d]{4}$', nhi):
            return False

        sum_letters = sum((str.find(alpha, str(nhi[i])) + 1) * (7-i) for i in range(3))

        sum_numbers = sum(int(nhi[i]) * (7-i) for i in range(3,6))

        total = sum_letters + sum_numbers
        check = 11 - (total % 11)

        if check == 11 or check % 10 != int(nhi[6]):
            return False

        return True


    @staticmethod
    def ref_num_pattern(ref_num):
        '''######-[SED/SBD] or SED-049-#### or SED17[NHI]'''
        ref_num = ref_num.strip().upper()
        if ref_num == '.':
            return False

        if re.fullmatch('^\d{6}[-|\s]?(SED|SDB|SED/SDB|SDB/SED|ACC)$', ref_num):
            return True
        
        if re.fullmatch('^SED-\d{3}-\d{4}$', ref_num):
            return True

        if re.fullmatch('^(SED|SDB)\d{2}[-|\s]?[A-H|J-N|P-Z]{3}\d{4}$', ref_num):
            return True

        return False


    def match_school(self, patient):
        for decile in deciles:
            for pattern in decile:
                if re.fullmatch(f"^{pattern}.*$", patient['school'].upper().strip()):
                    patient['decile'] = deciles[decile]
                    return True

        patient['decile'] = 0
        return False


    def _validate_patient(self, patient):
        '''Assumes self.needs_nhi and self.needs_ref'''
        flag = True

        if not self.nhi_checksum(patient['nhi']):
            flag = False
            self.needs_nhi.add(f"{patient['lname']}, {patient['fname']}")

        if not self.ref_num_pattern(patient['subid']):
            if patient['subid'] != '.':
                flag = False
                self.needs_ref.add(f"{patient['lname']}, {patient['fname']}")

        for key, item in patient.items():
            if not item and key != 'gender' and key != 'nhi':
                flag = False
                self.needs_other_info.add(f"{patient['lname']}, {patient['fname']}")

        return flag

    
    def validate_patients(self, patients):
        self.needs_nhi = set()
        self.needs_ref = set()
        self.needs_other_info = set()

        return tuple(filter(self._validate_patient, patients))


    def _validate_patient_OHSA(self, patient):
        flag = True

        if not self.nhi_checksum(patient['nhi']):
            flag = False
            self.needs_nhi.add(f"{patient['lname']}, {patient['fname']}")

        if not self.match_school(patient):
            flag = False
            self.needs_school.add(f"{patient['lname']}, {patient['fname']}")

        for key, item in patient.items():
            if not item and (key not in ('nhi', 'school', 'gender')):
                flag = False
                self.needs_other_info.add(f"{patient['lname']}, {patient['fname']}")

        return flag


    def validate_patients_OHSA(self, patients):
        self.needs_nhi = set()
        self.needs_school = set()
        self.needs_other_info = set()

        return tuple(filter(self._validate_patient_OHSA, patients))


    def _validate_patient_OHSA_PA(self, patient):
        flag = True

        if not self.nhi_checksum(patient['nhi']):
            flag = False
            self.needs_nhi.add(f"{patient['lname']}, {patient['fname']}")

        if not self.match_school(patient):
            flag = False
            self.needs_school.add(f"{patient['lname']}, {patient['fname']}")

        for key, item in patient.items():
            if not item and (key not in ('nhi', 'school', 'gender')):
                flag = False
                self.needs_other_info.add(f"{patient['lname']}, {patient['fname']}")

        if not patient['pa_num']:
            flag = False
            self.needs_prior_auth.add(f"{patient['lname']}, {patient['fname']}")

        return flag

    def validate_patients_OHSA_PA(self, patients):
        self.needs_nhi = set()
        self.needs_school = set()
        self.needs_other_info = set()
        self.needs_prior_auth = set()

        return tuple(filter(self._validate_patient_OHSA_PA, patients))



    def _validate_patient_PA(self, patient):
        flag = True

        if not patient['pa_num']:
            flag = False
            self.needs_prior_auth.add(f"{patient['lname']}, {patient['fname']}")

        if not self.nhi_checksum(patient['nhi']):
            flag = False
            self.needs_nhi.add(f"{patient['lname']}, {patient['fname']}")

        if not self.ref_num_pattern(patient['subid']):
            if patient['subid'] != '.':
                flag = False
            self.needs_ref.add(f"{patient['lname']}, {patient['fname']}")

        for key, item in patient.items():
            if not item and key != 'gender':
                flag = False
                self.needs_other_info.add(f"{patient['lname']}, {patient['fname']}")

        return flag


    def validate_patients_PA(self, patients):
        self.needs_nhi = set()
        self.needs_ref = set()
        self.needs_other_info = set()
        self.needs_prior_auth = set()

        return tuple(filter(self._validate_patient_PA, patients))


    def generate_claims_SDSC(self, patients):
        '''Generates a pdf of all patients and procedures on HP5959 claim forms, saves to U:/claims/SDSC/<year>/<month>/???'''
        pat_coords = {'nhi' : (140, 2985, {'charSpace' : 46}), 'birthdate' : (140, 2820, {'charSpace' : 47}), 'lname' : (1300, 2985, {}), 'fname' : (1300, 2820, {}),
            'address' : (1300, 2665, {}), 'city' : (1300, 2610, {}), 'school' : (140, 2460, {}), 'city2' : (140, 2290, {}), 'subid' : (1860, 1230, {})}
        proc_coords = {'qty' : (1300, 1150, {}), 'date' : (140, 1150, {}), 'fee' : (1860, 1150, {}), 'tooth' : (1500, 1150, {}), 'code' : (400, 1150, {})}

        c = canvas.Canvas(date.today().strftime("SDSC%d%m%y.pdf"), pagesize = (2480, 3508))

        for patient, procedures in patients:
            for page in range(((len(procedures) - 1) // 5) + 1):
                c.drawImage("U:\\Claims\\SDSC\\hp5959.jpg",0,0, width=2480, height=3508)
                c.setFont('Helvetica', 55)

                for i in pat_coords.keys():
                    c.drawString(*pat_coords[i][:2], patient[i], **pat_coords[i][2])
                
                subtotal = 0.0

                for num, i in enumerate(procedures[(5 * (page)):(5 * (page) + 5)]):
                    for j in proc_coords.keys():
                        c.drawString(proc_coords[j][0], proc_coords[j][1] - 74 * num, i[j], **proc_coords[j][2])

                    subtotal += float(i['fee'])

                c.drawRightString(2025, 325, f"{subtotal:.2f}")

                c.setLineWidth(7)
                if patient['subid'] == '.':
                    c.lines([(130,1780,160,1810), (130, 1810, 160, 1780)])
                else:
                    c.lines([(130,1980,160,2010), (130, 2010, 160, 1980)])

                if patient['gender'] == 0:
                    c.lines([(300, 2670 , 330 , 2700), (330, 2670, 300, 2700)])
                elif patient['gender'] == 1:
                    c.lines([(600, 2670 , 630 , 2700), (630, 2670, 600, 2700)])

                c.showPage()

        c.save()


    def generate_claims_PA(self, patients):
        '''Generates a pdf of all patients and procedures on HP5959 claim forms, saves to U:/claims/SDSC/<year>/<month>/???'''
        pat_coords = {'nhi' : (140, 2985, {'charSpace' : 46}), 'birthdate' : (140, 2820, {'charSpace' : 47}), 'lname' : (1300, 2985, {}), 'fname' : (1300, 2820, {}),
            'address' : (1300, 2665, {}), 'city' : (1300, 2610, {}), 'school' : (140, 2460, {}), 'city2' : (140, 2290, {}), 'subid' : (1860, 1230, {})}

        pa_coords = {'qty' : (1300, 680, {}), 'date' : (140, 680, {}), 'fee' : (1860, 680, {}), 'tooth' : (1500, 680, {}), 'code' : (400, 680, {})}

        c = canvas.Canvas(date.today().strftime("SDPA%d%m%y.pdf"), pagesize = (2480, 3508))

        for patient, procedures in patients:
            for page in range(((len(procedures) - 1) // 5) + 1):
                c.drawImage("U:\\Claims\\SDSC\\hp5959.jpg",0,0, width=2480, height=3508)
                # c.drawImage(".\\hp5959.jpg",0,0, width=2480, height=3508)
                c.setFont('Helvetica', 55)

                for i in pat_coords.keys():
                    c.drawString(*pat_coords[i][:2], patient[i], **pat_coords[i][2])
                
                c.drawString(1860, 770, patient['pa_num'])
                subtotal = 0.0

                for num, i in enumerate(procedures[(5 * (page)):(5 * (page) + 5)]):
                    for j in pa_coords.keys():
                        c.drawString(pa_coords[j][0], pa_coords[j][1] - 74 * num, i[j], **pa_coords[j][2])

                    subtotal += float(i['fee'])

                c.drawRightString(2025, 325, f"{subtotal:.2f}")

                c.setLineWidth(7)
                if patient['subid'] == '.':
                    c.lines([(130,1780,160,1810), (130, 1810, 160, 1780)])
                else:
                    c.lines([(130,1980,160,2010), (130, 2010, 160, 1980)])

                if patient['gender'] == 0:
                    c.lines([(300, 2670 , 330 , 2700), (330, 2670, 300, 2700)])
                elif patient['gender'] == 1:
                    c.lines([(600, 2670 , 630 , 2700), (630, 2670, 600, 2700)])

                c.showPage()

        c.save()


    def generate_claims_OHSA(self, patients):

        div = 70
        ohsa_proc_coords = {'date' : (150, 1290, {'charSpace' : 27}), 'code' : (485, 1290, {}), 'qty' : (1300, 1290, {}), 'tooth' : (1520, 1290, {}), 'fee' : (1900, 1290, {})}


        ohsa_pat_coords = {'nhi' : (150, 3020, {'charSpace' : 45}), 'birthdate' : (150, 2720, {'charSpace' : 49}), 'lname' : (150, 2870, {}),
                'fname' : (1300, 2870, {}), 'school' : (150, 2575, {}), 'decile' : (1300, 2575, {})}

        ohsa_cap_coords = { 'CON1' : {'date' : (150, 2300, {'charSpace' : 49}), 'box' : (975, 2325), 'fee' : (2100, 2300)},
                    'DBTOP1' : {'date' : (150, 1910, {'charSpace' : 27}), 'box' : (1070, 1940)},
                    'DBOPT1' : {'date' : (150, 1910-div, {'charSpace' : 27}), 'box' : (1070, 1940-div)},
                    'DBRAD1' : {'date' : (150, 1910-div*2, {'charSpace' : 27}), 'box' : (1070, 1940-div*2)},
                    'DBPBW1' : {'date' : (150, 1910-div*3, {'charSpace' : 27}), 'box' : (1070, 1940-div*3)},
                    'DBSCL1' : {'date' : (150, 1910-div*4, {'charSpace' : 27}), 'box' : (1070, 1940-div*4)},
                    'DBFIL1' : {'date' : (1300, 2000, {'charSpace' : 27}), 'box' : (2240, 2020), 'tooth' : (1850, 2000)},
                    'DBFIS1' : {'date' : (1300, 1760, {'charSpace' : 27}), 'box' : (2240, 1800), 'tooth' : (1820, 1780)}}

        c = canvas.Canvas(date.today().strftime("OHSA%d%m%y.pdf"),pagesize = (2480, 3508))

        tickbox = lambda x,y: c.lines([(x-15, y-15, x+15, y+15),(x-15, y+15, x+15, y-15)])

        for patient, procedures in patients:
            capitated = [procedure for procedure in procedures if procedure['code'] in ohsa_cap_coords.keys()]
            regular = [procedure for procedure in procedures if procedure['code'] not in ohsa_cap_coords.keys()]

            for page in range((((len(regular) - 1) // 8) + 1) if regular else 1):
                c.drawImage("U:\\Claims\\OHSA\\hp5953.jpg",0,0, width=2480, height=3508)
                c.setFont('Helvetica', 55)
                c.setLineWidth(7)

                for key in ohsa_pat_coords.keys():
                    c.drawString(ohsa_pat_coords[key][0], ohsa_pat_coords[key][1], str(patient[key]), **ohsa_pat_coords[key][2])

                if patient['gender'] == 0:
                    tickbox(1430, 2750)
                elif patient['gender'] == 1:
                    tickbox(1725, 2750)

                subtotal = 0.0

                if page == 0:
                    for proc in capitated:
                        c.drawString(ohsa_cap_coords[proc['code']]['date'][0], ohsa_cap_coords[proc['code']]['date'][1], str(proc['date']), **ohsa_cap_coords[proc['code']]['date'][2])
                        tickbox(*ohsa_cap_coords[proc['code']]['box'])
                        if proc['tooth']:
                            c.setFont('Helvetica', 35)
                            c.drawString(ohsa_cap_coords[proc['code']]['tooth'][0], ohsa_cap_coords[proc['code']]['tooth'][1], proc['tooth'])
                            c.setFont('Helvetica', 55)
                        if proc['fee'] :
                            c.drawRightString(ohsa_cap_coords[proc['code']]['fee'][0], ohsa_cap_coords[proc['code']]['fee'][1], proc['fee'])
                            subtotal += float(proc['fee'])

                for num, proc in enumerate(regular[page * 8:(page + 1) * 8]):
                    for field in proc.keys():
                        c.drawString(ohsa_proc_coords[field][0], ohsa_proc_coords[field][1] - (num*(div-2)), proc[field], **ohsa_proc_coords[field][2]) 
                    if proc['fee']:
                        subtotal += float(proc['fee'])

                c.drawRightString(2040, 200, f"{subtotal:.2f}")
                c.showPage()

        c.save()


    def generate_claims_OHSA_PA(self, patients):

        div = 70
        h = 625
        ohsa_proc_coords = {'date' : (150, h, {'charSpace' : 27}), 'code' : (485, h, {}), 'qty' : (1300, h, {}), 'tooth' : (1520, h, {}), 'fee' : (1900, h, {})}

        ohsa_pat_coords = {'nhi' : (150, 3020, {'charSpace' : 45}), 'birthdate' : (150, 2720, {'charSpace' : 49}), 'lname' : (150, 2870, {}),
                'fname' : (1300, 2870, {}), 'school' : (150, 2575, {}), 'decile' : (1300, 2575, {}), 'pa_num' : (1940, 720, {})}

        c = canvas.Canvas(date.today().strftime("OHPA%d%m%y.pdf"),pagesize = (2480, 3508))

        tickbox = lambda x,y: c.lines([(x-15, y-15, x+15, y+15),(x-15, y+15, x+15, y-15)])

        for patient, procedures in patients:

            for page in range((((len(procedures) - 1) // 6) + 1) if procedures else 1):
                c.drawImage("U:\\Claims\\OHSA\\hp5953.jpg",0,0, width=2480, height=3508)
                c.setFont('Helvetica', 55)
                c.setLineWidth(7)

                for key in ohsa_pat_coords.keys():
                    c.drawString(ohsa_pat_coords[key][0], ohsa_pat_coords[key][1], str(patient[key]), **ohsa_pat_coords[key][2])

                if patient['gender'] == 0:
                    tickbox(1430, 2750)
                elif patient['gender'] == 1:
                    tickbox(1725, 2750)

                subtotal = 0.0

                for num, proc in enumerate(procedures[page * 6:(page + 1) * 6]):
                    for field in proc.keys():
                        c.drawString(ohsa_proc_coords[field][0], ohsa_proc_coords[field][1] - (num*(div-2)), proc[field], **ohsa_proc_coords[field][2])
                    if proc['fee']:
                        subtotal += float(proc['fee'])

                c.drawRightString(2040, 200, f"{subtotal:.2f}")
                c.showPage()

        c.save()


    def create_claimpayment(self, amount, note, carrier):
        '''Creates an entry in the claimpayment table --> primay key for new entry'''
        self.sql(f'''INSERT INTO claimpayment (checkdate, checkamt, note, carriername, ClinicNum, DepositNum, IsPartial, PayType, SecUserNumEntry, Paygroup)
        VALUE (CURDATE(), {amount}, '{note}', '{carrier}', 0, 0, 0, 338, 11, 370)''') #de facto defaults

        self.sql('''select max(claimpaymentnum) from claimpayment''') #LAST_INSERT_ID() ?

        self.claim_payment_num = int(self.curs.fetchone()[0])


    # def get_sed_claims(self, month_ago):
    #     '''Set of sedation claims marked recieved from a given amount of months ago'''
    #     self.sql(f"""SELECT distinct(c.claimnum) FROM claim c INNER JOIN claimproc cp ON c.claimnum = cp.claimnum INNER JOIN procedurelog pl on cp.procnum = pl.procnum inner join procedurecode pc on pl.codenum = pc.codenum
    #         WHERE extract(year_month from pl.ProcDate) = extract(year_month from curdate() - interval {month_ago} month)
    #         and (pc.ProcCode = 'Sed-RA' OR pc.ProcCode = 'Sed-oral' OR pc.ProcCode = 'SED3' OR pc.ProcCode = 'FIS1')
    #         and c.PlanNum in (1, 14703) and c.ClaimStatus = 'U'""")
    #     return tuple(i[0] for i in self.curs.fetchall())

    @staticmethod
    def make_claim_insert_query(pat_details):
        #patnum, provnum, claimfee, inssubnum, inssubnum2
        insert_query = "insert into claim ({}) values ({})".format(', '.join(sed_insert_claim_dict.keys()), ', '.join(f"'{item}'" for item in sed_insert_claim_dict.values())).format(**pat_details)
        return insert_query


    @staticmethod
    def make_claimproc_insert_query(pat_details):
        #patnum, provnum, claimfee, inssubnum, inssubnum2
        insert_query = "insert into claimproc ({}) values ({})".format(''', '''.join(sed_insert_claimproc_dict.keys()), ''', '''.join(f"'{item}'" for item in sed_insert_claimproc_dict.values())).format(**pat_details)
        return insert_query


    def get_sed_procs(self):
        self.sql(SED_get_procedures.format(1))
        result = self.curs.fetchall()
        procs = tuple((patnum, tuple(procnums.split(','))) for (patnum, procnums) in result)
        print(*procs, sep='\n')
        return procs

    def get_sed_pats(self):
        self.sql(SED_get_patients.format(1))
        result = self.curs.fetchall()
        pats = tuple(f"{lname}, {fname}" for (lname, fname) in result)
        print(*pats, sep='\n')


    def monthly_sedation_report(self, month_ago = 1):
        sed_procs = self.get_sed_procs()
        #TODO


    def attach_claims(self, claims):
        '''Attaches claims to created entry in ins paymants'''
        assert self.claim_payment_num is not None, "claim_payment_num is not set"
        for claim in claims:
            self.sql(f"UPDATE claimproc SET claimpaymentnum = {self.claim_payment_num}, InsPayAmt = FeeBilled, status = 1, dateentry = CURDATE() WHERE claimnum = {claim}")
            self.sql(f"UPDATE claim SET claimstatus = 'R', DateReceived = CURDATE() WHERE claimnum = {claim}")


    def add_sent_today_CDA(self, carrier):
        '''Adds all claims sent today to sentClaim table; for SDSC and OHSA claims'''
        if carrier == 1:
            self.claimform = 32
            self.plan_num = 'in (1, 14703)'
            self.carrier = 'SDSC'
        elif carrier == 2:
            self.claimform = 33
            self.plan_num = '= 7'
            self.carrier = 'OHSA'
        elif carrier == 3:
            self.claimform = 21
            self.carrier = 'SDPA'
        elif carrier == 4:
            self.claimform = 34
            self.carrier = 'SED'
        elif carrier == 5:
            self.claimform = 35
            self.carrier = 'OHPA'

        self.sql(f"""SELECT claimnum FROM claim WHERE DateSent = CURDATE() and claimform  = {self.claimform} and ClaimStatus = 'S'""")
        claims = tuple(line[0] for line in self.curs.fetchall())

        for claim in claims:
            self.sql(f"""REPLACE INTO sentClaim (claimnum, claimset, carrier) value ({claim}, CONCAT('{self.carrier}', date_format(curdate(), '%d%m%y')), '{self.carrier}')""")


    def generate_spreadsheet(self):
        '''Generate .xlsx workbook to record claims sent in human readable format, assumes attributes: carrier, plan_num'''
        wb = Workbook()
        ws = wb.active

        ws['C1'] = date.today().strftime(f"{self.carrier}%d%m%y")
        ws.column_dimensions['C'].width = 35
        ws.column_dimensions['D'].width = 13
        ws.column_dimensions['E'].width = 13

        ws['A2'], ws['B2'], ws['C2'], ws['D2'], ws['E2'] = 'ClaimNum', 'NHI', 'Patient Name', 'Date', 'Fee'

        self.sql(f"""SELECT c.ClaimNum, p.SSN, CONCAT(p.LName, ', ', p.FName), DATE_FORMAT(c.DateService, '%d/%m/%Y'), c.ClaimFee
            FROM claim c INNER JOIN patient p ON c.PatNum = p.PatNum
            WHERE DateSent = CURDATE() AND ClaimStatus = 'S' AND claimform =  {self.claimform}
            ORDER BY p.LName, p.FName, c.DateService""")

        for index, row in enumerate(self.curs.fetchall()):
            for column, item in zip(['A', 'B', 'C', 'D', 'E'], row):
                ws[f'{column}{3 + index}'] = item
                if column == 'E':
                    ws[f'{column}{3 + index}'].style = 'Currency'

        num_rows = len(tuple(ws.rows))

        tab = Table(displayName="Table1", ref=f"A2:E{num_rows}")
        style = TableStyleInfo(name="TableStyleMedium1", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        tab.tableStyleInfo = style
        ws.add_table(tab)

        ws[f'E{num_rows + 1}'] = sum(ws[f'E{i}'].value for i in range(3, num_rows + 1))
        ws[f'E{num_rows + 2}'] = ws[f'E{num_rows + 1}'].value * 0.15
        ws[f'E{num_rows + 3}'] = ws[f'E{num_rows + 1}'].value * 1.15

        for i in range(1,4):
            ws[f'E{num_rows + i}'].style = 'Currency'

        self.metadata = [num_rows - 2] + [f"{(ws[f'E{num_rows + i}'].value):.2f}" for i in range(1, 4)]

        wb.save(date.today().strftime(f"{self.carrier}%d%m%y.xlsx"))


    def generate_cover(self):
        '''Generate CDA coversheet as .docx, assumes attributes: carrier, metadata'''
        if self.carrier in ('SDSC', 'SDPA'):
            doc = DocxTemplate("S:\\test folder please ignore\\Python stuff\\sed_returns\\ayo.docx")
        elif self.carrier in ('OHSA', 'OHPA'):
            doc = DocxTemplate("S:\\test folder please ignore\\Python stuff\\sed_returns\\ayo1.docx")

        context_keys = ['r1', 'r2', 'r3', 'r4', 'r5', 'r6', 'r7', 'r8', 'r9', 'r10', 'num_pat', 'total_ex', 'gst', 'total_inc']
        context = dict()

        for key, char in zip(context_keys[:10], date.today().strftime(f"{self.carrier}%d%m%y")):
            context[key] = str(char)

        for key, char in zip(context_keys[10:], self.metadata):
            context[key] = str(char)

        doc.render(context)
        # os.chdir('.\\Summary Sheets')
        doc.save(date.today().strftime(f"{self.carrier}%d%m%y.docx"))


    def print_paperwork(self):
        '''Prints claim summary and cover form, assumes attribute: carrier'''
        with in_dir('.\\Summary Sheets'):
            os.startfile(date.today().strftime(f"{self.carrier}%d%m%y.docx"), "print")
            sleep(3)
            os.startfile(date.today().strftime(f"{self.carrier}%d%m%y.docx"), "print")

        os.startfile(date.today().strftime(f"{self.carrier}%d%m%y.xlsx"), "print")

        with in_dir(".\\woo"):
            os.startfile(date.today().strftime(f"{self.carrier}%d%m%y.pdf"), "print")


    def print_paperwork_test(self):
        '''Prints claim summary and cover form, assumes attribute: carrier'''
        os.startfile(date.today().strftime(f"{self.carrier}%d%m%y.docx"), "print")
        sleep(3)
        os.startfile(date.today().strftime(f"{self.carrier}%d%m%y.docx"), "print")

        os.startfile(date.today().strftime(f"{self.carrier}%d%m%y.xlsx"), "print")

        os.startfile(date.today().strftime(f"{self.carrier}%d%m%y.pdf"), "print")


    def received_claim(self, claim_set, moh_code, amount):
        '''Updates sentClaim table, creates andattaches claims to insurance payment'''
        self.sql(f"SELECT claimnum FROM sentClaim WHERE claimSet = '{claim_set}'")
        result = self.curs.fetchall()
        if result:
            claims = tuple(line[0] for line in result)
        else:
            return 0
        # self.sql(f"DELETE FROM sentClaim WHERE claimSet = '{claim_set}'")
        self.sql(f"UPDATE sentClaim SET status = 1 WHERE claimSet = '{claim_set}'")

        carrier = ''.join(i for i in claim_set if i.isalpha())
        if carrier == 'W':
            carrier = 'ACC'
        elif carrier == 'SED':
            carrier = 'SDSC'

        self.create_claimpayment(amount, f"{claim_set} // {moh_code}", carrier)
        self.attach_claims(claims)

        return 1


    def received_ACC(self):
        '''Gets a list of ACC claim claimnums to be fed into received_claim method''' 
        claims = set()

        print("Once all numbers have been added, enter 'done' instead")

        entry = input('Enter invoice number: ')
        while entry.strip().lower() != 'done':
            self.sql(f"SELECT claimnum FROM sentClaim WHERE claimSet = '{entry.strip().upper()}'")
            try:
                claims.add(self.curs.fetchone()[0])
            except TypeError:
                print('Invoice number not found')
            entry = input('Enter invoice number: ')

        if claims:
            for claim in claims:
                self.sql(f"UPDATE sentClaim SET status = 1 WHERE claimnum = '{claim}'")

            note = input("Payment reference: ")
            self.create_claimpayment(input("Amount: "), f"ACC{note}", "ACC")
            self.attach_claims(claims)
        else:
            print('No claims were added')
            return


    def cleanup_reject(self, claim):
        '''Clears an item from sentClaim table for when it is completely rejected'''
        self.sql(f"select * from sentclaim where claimnum = {claim}")
        if self.curs.fetchone():
            self.sql(f"delete from sentclaim where claimnum = {claim}")
            self.sql(f"update claim set claimstatus = 'W' where claimnum = {claim}")
            return 1
        else:
            return 0

    def set_as_sent(self, claims):
        for claim in claims:
            self.sql(set_as_sent.format(claim))


    def SDSC_info(self):
        return("{} ready to send\n".format(len(self.validate_patients(self.get_patients_SDSC()))))


    def OHSA_info(self):
        return("{} ready to send\n".format(len(self.validate_patients_OHSA(self.get_patients_OHSA()))))

        # if sc.needs_nhi:
        #     print("Needs NHI:", end='\n\t')
        #     print(*sc.needs_nhi, sep='\n\t')

        # if sc.needs_ref:
        #     print("Needs sed-ref:", end='\n\t')    
        #     print(*sc.needs_ref, sep='\n\t')

        # if sc.needs_other_info:
        #     print("Needs other info:", end='\n\t')
        #     print(*sc.needs_other_info, sep='\n\t')


    def send_SDSC(self):
        patients = self.validate_patients(self.get_patients_SDSC())[:50]
        self.set_as_sent([i['claimnum'] for i in patients])
        self.add_sent_today_CDA(1)

        with in_dir(date.today().strftime(f'U:\\Claims\\SDSC\\%Y\\%m - %b\\{self.carrier}%d%m%y')):

            self.generate_claims_SDSC(zip(patients, tuple(self.get_procedures(patient) for patient in patients)))

            self.generate_spreadsheet()

            self.generate_cover()

            # self.print_paperwork_test()


    def send_OHSA(self):
        patients = self.validate_patients_OHSA(self.get_patients_OHSA())[:50]
        self.set_as_sent([i['claimnum'] for i in patients])
        self.add_sent_today_CDA(2)

        with in_dir(date.today().strftime(f'U:\\Claims\\OHSA\\%Y\\%m - %b\\{self.carrier}%d%m%y')):
        # with in_dir('S:\\test folder please ignore\\Python stuff'):

            self.generate_claims_OHSA(zip(patients, tuple(self.get_procedures_OHSA(patient) for patient in patients)))

            self.generate_spreadsheet()

            self.generate_cover()

            # self.print_paperwork_test()


    def send_OHSA_PA(self):
        patients = self.validate_patients_OHSA_PA(self.get_patients_OHSA_PA())
        self.set_as_sent([i['claimnum'] for i in patients])
        self.add_sent_today_CDA(5)

        with in_dir(date.today().strftime(f'U:\\Claims\\OHSA\\%Y\\%m - %b\\{self.carrier}%d%m%y')):

            self.generate_claims_OHSA_PA(zip(patients, tuple(self.get_procedures_OHSA(patient) for patient in patients)))

            self.generate_spreadsheet()

            self.generate_cover()

            # self.print_paperwork_test()

    def OHPA_test(self):
        with in_dir('S:\\test folder please ignore\\Python stuff'):
            patients = self.validate_patients_OHSA_PA(self.get_patients_OHSA_PA())
            self.generate_claims_OHSA_PA(zip(patients, tuple(self.get_procedures_OHSA(patient) for patient in patients)))
            # print(*zip(patients, tuple(self.get_procedures_OHSA(patient) for patient in patients)))



    def send_PA(self):
        patients = self.validate_patients_PA(self.get_patients_PA())
        self.set_as_sent([i['claimnum'] for i in patients])
        self.add_sent_today_CDA(3)

        with in_dir(date.today().strftime(f'U:\\Claims\\SDSC\\%Y\\%m - %b\\{self.carrier}%d%m%y')):

            self.generate_claims_PA(zip(patients, tuple(self.get_procedures_PA(patient) for patient in patients)))

            self.generate_spreadsheet()

            self.generate_cover()

            # self.print_paperwork_test()


    def send_SED(self):
        self.sql("""SELECT p.SSN, CONCAT(p.lname, ', ', p.fname), ins.subscriberid, pc.descript, DATE_FORMAT(pl.procdate, '%d/%m/%y'), pl.procfee
            FROM claim c
            INNER JOIN claimproc cp on c.claimnum = cp.claimnum
            INNER JOIN procedurelog pl on cp.procnum = pl.procnum
            INNER JOIN inssub ins on cp.inssubnum = ins.inssubnum
            INNER JOIN procedurecode pc on pl.codenum = pc.codenum
            INNER JOIN patient p on pl.patnum = p.patnum
            WHERE c.claimform = 34 and year(c.dateservice) = 2018 and month(c.dateservice) = 4 and c.claimstatus = 'W'
            """)
        result = self.curs.fetchall()
        #etc etc





class DashBoard(SendClaims):

    cfg = {'padx' : 10, 'pady' : 10}

    def __init__(self):
        super().__init__()

        self.root = tk.Tk()
        self.root.title("Claiming Dashboard")

        h = 220
        w = 450
        self.root.minsize(height=h, width=w)
        self.root.maxsize(height=h, width=w)

        self.frame = tk.Frame(self.root)

        self.footer = tk.Frame(self.root)

        self.home()

    def __call__(self):
        '''wrapper for style points'''
        self.root.mainloop()


    def make_frame(func):
        def wrap(self):
            for item in self.frame.winfo_children():
                item.destroy()

            self.frame.pack()

            for item in self.footer.winfo_children():
                item.destroy()

            self.footer.pack()

            tk.Button(self.footer, text='Back', command=self.home).grid(**self.cfg)

            func(self)

        return wrap


    @make_frame
    def home(self):
        tk.Button(self.frame, height=5, width=15, text='Send Claims', command=self.send_frame).grid(row = 0,column = 0, **self.cfg)
        tk.Button(self.frame, height=5, width=15, text='Receive Claims', command=self.recv_frame).grid(row = 0,column = 1, **self.cfg)

        for item in self.footer.winfo_children():
            item.destroy()

        tk.Button(self.footer, text='Exit', command=self.root.quit).grid(**self.cfg)

    @make_frame
    def send_frame(self):
        tk.Button(self.frame, height=5, width=15, text='SDSC', command=self.send_SDSC_frame).grid(row=0, column = 0, **self.cfg)
        tk.Button(self.frame, height=5, width=15, text='OHSA', command=self.send_OHSA_frame).grid(row=0, column = 1, **self.cfg)
        tk.Button(self.frame, height=5, width=15, text='ACC', command=self.send_ACC_frame).grid(row=0, column = 2, **self.cfg)


    @make_frame
    def send_SDSC_frame(self):
        tk.Label(self.frame, text=self.SDSC_info()).grid(row=0, column=0, **self.cfg)
        tk.Button(self.frame, height=5, width=15, text='Go', command=self.send_SDSC_action).grid(row=0, column=1, **self.cfg)


    def send_SDSC_action(self):
        self.send_SDSC()
        self.send_SDSC_frame()


    @make_frame
    def send_OHSA_frame(self):
        '''test'''
        tk.Label(self.frame, text=self.OHSA_info()).grid(**self.cfg)
        tk.Button(self.frame, height=5, width=15, text='Go', command=self.send_OHSA_action).grid(row=0, column=1, **self.cfg)


    def send_OHSA_action(self):
        self.send_OHSA()
        self.send_OHSA_frame()


    @make_frame
    def send_ACC_frame(self):
        '''It just launches the old thing'''
        tk.Button(self.frame, height=5, width=15, text='Go', command=self.send_ACC_action).grid()


    def send_ACC_action(self):
        '''TODO: make better'''
        os.startfile('acc_automater.py')


    @make_frame
    def recv_frame(self):
        tk.Button(self.frame, height=5, width=15, text='SDSC', command=self.recv_SDSC_frame).grid(row=0, column = 0, **self.cfg)
        tk.Button(self.frame, height=5, width=15, text='OHSA', command=self.recv_OHSA_frame).grid(row=0, column = 1, **self.cfg)
        tk.Button(self.frame, height=5, width=15, text='ACC',  command=self.recv_ACC_frame).grid(row=0, column = 2, **self.cfg)

        tk.Button(self.frame, text='Clear rejected', command = self.clear_reject_frame).grid(row=1, column=0, columnspan=3)


    @make_frame
    def clear_reject_frame(self):
        tk.Label(self.frame, text="Enter claimnum").grid(row=0, column=0, sticky='E')
        
        self.stvar = tk.StringVar()
        tk.Entry(self.frame, textvariable=self.stvar).grid(row=0, column=1)

        tk.Button(self.frame, text='Go', command=self.clear_reject_action).grid(row=1, column=0)
        self.lab = tk.Label(self.frame, text='')
        self.lab.grid(row=1, column=1)


    def clear_reject_action(self):
        if self.cleanup_reject(self.stvar.get()):
            self.lab.config(text = 'Success!')
        else:
            self.lab.config(text = 'Could not find claim')
        self.stvar.set('')



    @make_frame
    def recv_SDSC_frame(self):
        tk.Label(self.frame, text="Our ref").grid(row=0, column=0, sticky='E')
        tk.Label(self.frame, text="MoH ref").grid(row=1, column=0, sticky='E')
        tk.Label(self.frame, text="Amount").grid(row=2, column=0, sticky='E')

        self.res = tk.StringVar()
        tk.Label(self.frame, textvariable=self.res).grid(row=3, column=0)

        self.sv = tuple(tk.StringVar() for _ in range(3))

        for i, v in enumerate(self.sv):        
            tk.Entry(self.frame, textvariable=v).grid(row=i, column=1)

        tk.Button(self.frame, text='Go', command=self.recv_SDSC, width=15, height=5).grid(row=0, column=2, rowspan=3, padx=15)


    def recv_SDSC(self):
        self.res.set('Working...')

        if self.received_claim(*[v.get() for v in self.sv]):
            self.res.set('Success!')
        else:
            self.res.set('Could not find claim')

        for v in self.sv:
            v.set('')


    @make_frame
    def recv_OHSA_frame(self):
        tk.Label(self.frame, text="Our ref").grid(row=0, column=0, sticky='E')
        tk.Label(self.frame, text="MoH ref").grid(row=1, column=0, sticky='E')
        tk.Label(self.frame, text="Amount").grid(row=2, column=0, sticky='E')

        self.res = tk.StringVar()
        tk.Label(self.frame, textvariable=self.res).grid(row=3, column=0)

        self.sv = tuple(tk.StringVar() for _ in range(3))

        for i, v in enumerate(self.sv):        
            tk.Entry(self.frame, textvariable=v).grid(row=i, column=1)

        tk.Button(self.frame, text='Go', command=self.recv_OHSA, width=15, height=5).grid(row=0, column=2, rowspan=3, padx=15)


    def recv_OHSA(self):
        self.res.set('Working...')

        if self.received_claim(*[v.get() for v in self.sv]):
            self.res.set('Success!')
        else:
            self.res.set('Could not find claim')

        for v in self.sv:
            v.set('')


    @make_frame
    def recv_ACC_frame(self):
        self.acc_claim_info = tuple(tk.StringVar() for _ in range(2))
        self.sv = tuple(tk.StringVar() for _ in range(2))
        self.claims = set()
        self.adder = self.ACC_add_claim()
        self.adder.send(None)

        tk.Label(self.frame, text="Payment ref").grid(row=0, column=0, sticky='E', **self.cfg)
        tk.Label(self.frame, text="Amount").grid(row=1, column=0, sticky='E', **self.cfg)

        tk.Entry(self.frame, textvariable=self.acc_claim_info[0]).grid(row=0, column=1, **self.cfg)
        tk.Entry(self.frame, textvariable=self.acc_claim_info[1]).grid(row=1, column=1, **self.cfg)

        tk.Label(self.frame, text="Enter invoice #:").grid(row=2, column=0, sticky='E', **self.cfg)
        tk.Entry(self.frame, textvariable=self.sv[0]).grid(row=2, column=1, **self.cfg)
        tk.Button(self.frame, text='Add', command=self._acc_adder).grid(row=2, column=2, sticky='W', **self.cfg)

        tk.Label(self.frame, textvariable=self.sv[1]).grid(row=3, column=0, **self.cfg)
        tk.Button(self.frame, text='Done', command=self.recv_ACC_action).grid(row=3, column=2, **self.cfg)


    def _acc_adder(self):
        self.adder.send(self.sv[0].get().strip().upper())
        # self.sv[0].set('')


    def ACC_add_claim(self):
        while True:
            self.sql(f"SELECT claimnum FROM sentclaim WHERE claimset = '{(yield)}' and CARRIER = 'ACC'")
            result = self.curs.fetchone()
            if result:
                self.claims.add(result[0])
                self.sv[1].set(f'{self.sv[0].get()} Added!')
            else:
                self.sv[1].set('Could not find claim')


    def recv_ACC_action(self):
        if not self.claims:
            self.sv[1].set('No claims were added')
            return

        for claim in self.claims:
            self.sql(f"UPDATE sentClaim SET status = 1 WHERE claimnum = '{claim}'")

        self.create_claimpayment(self.acc_claim_info[1].get(), f"ACC {self.acc_claim_info[0].get()}", "ACC")
        self.attach_claims(self.claims)

        self.sv[1].set('Success!')

        self.claims.clear() #important lol

        self.acc_claim_info[0].set('')
        self.acc_claim_info[1].set('')
        self.sv[0].set('')



if __name__ == '__main__':
    with DashBoard() as db:
        db()
